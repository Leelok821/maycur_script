# -*- coding:utf-8 -*-

"""
# Author ：li zi hao
"""

import requests
import openpyxl


def get_data():
    wb = openpyxl.load_workbook('./万旗服饰需批量重置 密码人员.xlsx')
    ws = wb['结果']
    list = []
    for i in ws:
        list2 = []
        for data in i:
            list2.append(data.value)
        list.append(list2)
    return list
    wb.close()


def write_data(write_list):
    wb = openpyxl.load_workbook('./万旗服饰需批量重置 密码人员.xlsx')
    ws = wb.worksheets[0]
    for data in write_list:
        ws.cell(data[0], data[1], data[2])
    wb.save('万旗服饰需批量重置 密码人员.xlsx')


def do_reset():
    data = get_data()[1:]
    write_list = []
    for i in data:
        index = data.index(i)
        user_code = i[2]
        url = f'https://boss.maycur.com/backend/enterprise/employee/EC2106221KEF8KH7/{user_code}'
        header = {
            "tokenId": "eyJhbGciOiJIUzI1NiJ9.eyJqdGkiOiJtYXljdXJfand0X2lkIiwic3ViIjoiVUkyMTA4MzEyMldJTDhIVyIsImlhdCI6MTY0ODY5MDgyMiwiYXVkIjoiQk9TUyIsImV4cCI6MTY0ODY5ODAyMn0.tKd8bu6CW47y862iujywyVb0B_5Rf1i28-Qw4cKCAAA",
            "user": "UI21083122WIL8HW"
        }
        method = 'put'
        res = requests.request(method=method, url=url, headers=header).json()
        password = res.get('payload')
        print(f'usercode是{user_code},pwd是{password}')
        write_list.append((index+2,9,password))
    write_data(write_list)






do_reset()
